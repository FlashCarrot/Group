from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, jsonify
import os
from werkzeug.utils import secure_filename
import csv
from database import OracleDB
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import (
    PieChart, ProjectedPieChart, Reference,
    BarChart, Series, LineChart
)
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd
import plotly.express as px
import plotly.utils
import json

######################################################################


app = Flask(__name__)
app.config['REPORT_FOLDER'] = 'reports'
app.config['UPLOAD_DIRECTORY'] = 'uploads'
app.secret_key = 'byebye'


# Home Page ##########################################################

@app.route('/')
@app.route('/home')
def home():
    return render_template('home.html')


# Data Page #########################################################
@app.route("/group_data.html", methods=["Get", "POST"])
def group_data():
    if request.method == 'POST':
        if 'datafile' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
        file = request.files['datafile']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)

        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_DIRECTORY'], filename)
        file.save(file_path)

        data = []
        with open(file_path) as file_object:
            reader_obj = csv.reader(file_object)
            next(reader_obj)  # Skip header row

            for row in reader_obj:
                # Remove user_id from data processing
                data.append([
                    row[0],  # age
                    row[1],  # gender
                    row[2],  # technology_usage_hours
                    row[3],  # social_media_usage_hours
                    row[4],  # gaming_hours
                    row[5],  # screen_time_hours
                    row[6],  # mental_health_status
                    row[7],  # stress_level
                    row[8],  # sleep_hours
                    row[9],  # physical_activity_hours
                    row[10], # support_systems_access
                    row[11], # work_environment_impact
                    row[12]  # online_support_usage
                ])

        if data:
            with OracleDB().get_connection() as connection:
                insert_statement = """
                    INSERT INTO FY_GROUP_DATA
                    (age,
                    gender,
                    technology_usage_hours,
                    social_media_usage_hours,
                    gaming_hours,
                    screen_time_hours,
                    mental_health_status,
                    stress_level,
                    sleep_hours,
                    physical_activity_hours,
                    support_systems_access,
                    work_environment_impact,
                    online_support_usage)
                    VALUES (:1, :2, :3, :4, :5, :6, :7, :8, :9, :10, :11, :12, :13)
                """
                cursor = connection.cursor()
                cursor.executemany(insert_statement, data)
                connection.commit()

    with OracleDB().get_connection() as connection:
        query = '''
                SELECT * FROM fy_group_data
                ORDER BY data_pk
        '''
        cursor = connection.cursor()
        cursor.execute(query)
        data = cursor.fetchall()

    return render_template("group_data.html", data=data)

# Data Edit ############################################################

@app.route('/group_edit/<id>', methods=['GET', 'POST'])
def group_edit(id):
    if request.method == "GET":
        with OracleDB().get_connection() as connection:
            cursor = connection.cursor()
            cursor.execute("""
                SELECT * FROM FY_GROUP_DATA 
                WHERE DATA_PK = :id
            """, {'id': id})
            data = cursor.fetchone()
            print("Debug - Edit Data:", data)  # Add this debug line
            return render_template('group_edit.html', data=data)

# Data Add ############################################################

@app.route('/group_add', methods=['GET', 'POST'])
def group_add():
    if request.method == "POST":
        # Get form data
        age = request.form.get("AGE")
        gender = request.form.get("GENDER")
        tech_hours = request.form.get("TECHNOLOGY_USAGE_HOURS")
        social_hours = request.form.get("SOCIAL_MEDIA_USAGE_HOURS")
        gaming_hours = request.form.get("GAMING_HOURS")
        screen_hours = request.form.get("SCREEN_TIME_HOURS")
        mental_status = request.form.get("MENTAL_HEALTH_STATUS")
        stress_level = request.form.get("STRESS_LEVEL")
        sleep_hours = request.form.get("SLEEP_HOURS")
        physical_hours = request.form.get("PHYSICAL_ACTIVITY_HOURS")
        support_access = request.form.get("SUPPORT_SYSTEMS_ACCESS")
        work_impact = request.form.get("WORK_ENVIRONMENT_IMPACT")
        online_support = request.form.get("ONLINE_SUPPORT_USAGE")

        with OracleDB().get_connection() as connection:
            cursor = connection.cursor()
            insert_statement = """
                INSERT INTO FY_GROUP_DATA (
                    AGE,
                    GENDER,
                    TECHNOLOGY_USAGE_HOURS,
                    SOCIAL_MEDIA_USAGE_HOURS,
                    GAMING_HOURS,
                    SCREEN_TIME_HOURS,
                    MENTAL_HEALTH_STATUS,
                    STRESS_LEVEL,
                    SLEEP_HOURS,
                    PHYSICAL_ACTIVITY_HOURS,
                    SUPPORT_SYSTEMS_ACCESS,
                    WORK_ENVIRONMENT_IMPACT,
                    ONLINE_SUPPORT_USAGE
                ) VALUES (
                    :1, :2, :3, :4, :5, :6, :7, :8, :9, :10, :11, :12, :13
                )
            """
            cursor.execute(insert_statement, [
                age, gender, tech_hours, social_hours, gaming_hours,
                screen_hours, mental_status, stress_level, sleep_hours,
                physical_hours, support_access, work_impact, online_support
            ])
            connection.commit()
            return redirect(url_for('group_data'))

    return render_template('group_add.html', title="Add Item")


# Data Delete ###################################################

@app.route('/group_delete/<id>', methods=['GET', 'POST'])
def group_delete(id):
    data = None
    print("in group_delete:", id)
    with OracleDB().get_connection() as connection:
        cursor = connection.cursor()
        sql = "delete from FY_GROUP_DATA where data_pk = :data_pk"
        cursor.execute(sql, {'data_pk': id})
        connection.commit()
    return redirect(url_for('group_data'))


# Data Download ###################################################

@app.route('/group_download')
def group_download():
    wb = Workbook()
    
    # Data Sheet (existing sheet) ##########################
    ws_data = wb.active
    ws_data.title = "Raw Data"

    with OracleDB().get_connection() as connection:
        # Original data query
        query = "SELECT * FROM fy_group_data ORDER BY data_pk"
        cursor = connection.cursor()
        cursor.execute(query)
        data = cursor.fetchall()

        # Headers for raw data
        headers = ["Data ID", "User ID", "Age", "Gender", "Technology Usage Hours", 
                  "Social Media Usage Hours", "Gaming Hours", "Screen Time Hours",
                  "Mental Health Status", "Stress Level", "Sleep Hours", 
                  "Physical Activity Hours", "Support Systems Access",
                  "Work Environment Impact", "Online Support Usage"]
        
        for col, header in enumerate(headers, 1):
            ws_data.cell(row=1, column=col).value = header
            ws_data.cell(row=1, column=col).style = "Accent1"

        # Data rows
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws_data.cell(row=row_idx, column=col_idx).value = value

        # Mental Health Analysis Sheet ##########################
        ws_mental = wb.create_sheet("Mental Health Analysis")
        cursor.execute("""
            SELECT 
                mental_health_status,
                COUNT(*) as count,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER (), 1) as percentage
            FROM fy_group_data 
            WHERE technology_usage_hours >= 8
            AND mental_health_status IS NOT NULL
            GROUP BY mental_health_status
            ORDER BY count DESC
        """)
        mental_health_data = cursor.fetchall()

        ws_mental.append(["Mental Health Distribution Among Heavy Tech Users (>8hrs)"])
        ws_mental.append(["Mental Health Status", "Count", "Percentage"])
        for row in mental_health_data:
            ws_mental.append(row)

        # Sleep Analysis Sheet ##########################
        ws_sleep = wb.create_sheet("Sleep Analysis")
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN screen_time_hours < 4 THEN 'Low Screen Time (< 4hrs)'
                    WHEN screen_time_hours < 8 THEN 'Medium Screen Time (4-8hrs)'
                    ELSE 'High Screen Time (> 8hrs)'
                END as screen_time_level,
                ROUND(AVG(sleep_hours), 2) as avg_sleep_hours,
                COUNT(*) as sample_size
            FROM fy_group_data
            WHERE screen_time_hours IS NOT NULL
            AND sleep_hours IS NOT NULL
            GROUP BY 
                CASE 
                    WHEN screen_time_hours < 4 THEN 'Low Screen Time (< 4hrs)'
                    WHEN screen_time_hours < 8 THEN 'Medium Screen Time (4-8hrs)'
                    ELSE 'High Screen Time (> 8hrs)'
                END
            ORDER BY avg_sleep_hours DESC
        """)
        sleep_data = cursor.fetchall()

        ws_sleep.append(["Screen Time vs Average Sleep Hours"])
        ws_sleep.append(["Screen Time Level", "Average Sleep Hours", "Sample Size"])
        for row in sleep_data:
            ws_sleep.append(row)

        # Stress Analysis Sheet ##########################
        ws_stress = wb.create_sheet("Stress Analysis")
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN social_media_usage_hours < 2 THEN '< 2hrs'
                    WHEN social_media_usage_hours < 4 THEN '2-4hrs'
                    WHEN social_media_usage_hours < 6 THEN '4-6hrs'
                    ELSE '> 6hrs'
                END as social_media_usage,
                stress_level,
                COUNT(*) as count
            FROM fy_group_data
            WHERE social_media_usage_hours IS NOT NULL
            AND stress_level IS NOT NULL
            GROUP BY 
                CASE 
                    WHEN social_media_usage_hours < 2 THEN '< 2hrs'
                    WHEN social_media_usage_hours < 4 THEN '2-4hrs'
                    WHEN social_media_usage_hours < 6 THEN '4-6hrs'
                    ELSE '> 6hrs'
                END,
                stress_level
            ORDER BY social_media_usage
        """)
        stress_data = cursor.fetchall()

        ws_stress.append(["Social Media Usage vs Stress Levels"])
        ws_stress.append(["Social Media Usage", "Stress Level", "Count"])
        for row in stress_data:
            ws_stress.append(row)

        # Physical Activity Analysis Sheet ##########################
        ws_activity = wb.create_sheet("Physical Activity Analysis")
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN gaming_hours < 2 THEN 'Light Gamer (< 2hrs)'
                    WHEN gaming_hours < 4 THEN 'Moderate Gamer (2-4hrs)'
                    ELSE 'Heavy Gamer (> 4hrs)'
                END as gaming_level,
                ROUND(AVG(physical_activity_hours), 2) as avg_physical_activity,
                COUNT(*) as sample_size
            FROM fy_group_data
            WHERE gaming_hours IS NOT NULL
            AND physical_activity_hours IS NOT NULL
            GROUP BY 
                CASE 
                    WHEN gaming_hours < 2 THEN 'Light Gamer (< 2hrs)'
                    WHEN gaming_hours < 4 THEN 'Moderate Gamer (2-4hrs)'
                    ELSE 'Heavy Gamer (> 4hrs)'
                END
            ORDER BY avg_physical_activity DESC
        """)
        activity_data = cursor.fetchall()

        ws_activity.append(["Gaming Hours vs Physical Activity"])
        ws_activity.append(["Gaming Level", "Average Physical Activity Hours", "Sample Size"])
        for row in activity_data:
            ws_activity.append(row)

    # Add charts to analysis sheets
    # 1. Mental Health Pie Chart
    pie = PieChart()
    labels = Reference(ws_mental, min_col=1, min_row=3, max_row=2+len(mental_health_data))
    data = Reference(ws_mental, min_col=2, min_row=2, max_row=2+len(mental_health_data))
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = "Mental Health Distribution"
    ws_mental.add_chart(pie, "E2")

    # 2. Sleep Analysis Bar Chart
    bar1 = BarChart()
    bar1.type = "col"
    bar1.style = 10
    labels = Reference(ws_sleep, min_col=1, min_row=3, max_row=2+len(sleep_data))
    data = Reference(ws_sleep, min_col=2, min_row=2, max_row=2+len(sleep_data))
    bar1.add_data(data, titles_from_data=True)
    bar1.set_categories(labels)
    bar1.title = "Average Sleep Hours by Screen Time"
    bar1.y_axis.title = "Hours"
    ws_sleep.add_chart(bar1, "E2")

    # 3. Stress Analysis Stacked Bar Chart
    bar2 = BarChart()
    bar2.type = "col"
    bar2.style = 10
    labels = Reference(ws_stress, min_col=1, min_row=3, max_row=2+len(stress_data))
    data = Reference(ws_stress, min_col=3, min_row=2, max_row=2+len(stress_data))
    bar2.add_data(data, titles_from_data=True)
    bar2.set_categories(labels)
    bar2.title = "Stress Levels by Social Media Usage"
    bar2.y_axis.title = "Count"
    ws_stress.add_chart(bar2, "E2")

    # 4. Physical Activity Line Chart
    line = BarChart()
    line.type = "col"
    line.style = 10
    labels = Reference(ws_activity, min_col=1, min_row=3, max_row=2+len(activity_data))
    data = Reference(ws_activity, min_col=2, min_row=2, max_row=2+len(activity_data))
    line.add_data(data, titles_from_data=True)
    line.set_categories(labels)
    line.title = "Average Physical Activity by Gaming Level"
    line.y_axis.title = "Hours"
    ws_activity.add_chart(line, "E2")

    # Adjust the size of charts
    for ws in [ws_mental, ws_sleep, ws_stress, ws_activity]:
        for chart in ws._charts:
            chart.width = 15  # Width in centimeters
            chart.height = 10  # Height in centimeters

    # Formatting and styling for all sheets
    for ws in [ws_data, ws_mental, ws_sleep, ws_stress, ws_activity]:
        # Style the headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # For analysis sheets, style the title and merge cells
        if ws != ws_data:
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:C1')
            ws['A1'].alignment = Alignment(horizontal='center')

        # Adjust column widths - Modified to handle merged cells
        for column_cells in ws.columns:
            length = max(len(str(cell.value) if cell.value is not None else "") 
                        for cell in column_cells)
            if length > 0:
                column_letter = get_column_letter(column_cells[0].column)
                adjusted_width = (length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    filename = "digital_wellbeing_analysis.xlsx"
    wb.save(app.config['REPORT_FOLDER'] + "/" + filename)

    return send_from_directory(app.config['REPORT_FOLDER'], filename, as_attachment=True)


# Dashboard ############################################################
@app.route('/dashboard')
def dashboard():
    with OracleDB().get_connection() as connection:
        cursor = connection.cursor()
        
        # Check if there's data first
        cursor.execute("SELECT COUNT(*) FROM FY_GROUP_DATA")
        if cursor.fetchone()[0] == 0:
            flash('No data available. Please upload or add data first.', 'warning')
            return redirect(url_for('group_data'))
            
        # 1. Mental Health Distribution for Heavy Tech Users
        cursor.execute("""
            SELECT 
                mental_health_status,
                COUNT(*) as count,
                ROUND(COUNT(*) * 100.0 / SUM(COUNT(*)) OVER (), 1) as percentage
            FROM fy_group_data 
            WHERE technology_usage_hours >= 8
            AND mental_health_status IS NOT NULL
            GROUP BY mental_health_status
            ORDER BY count DESC
        """)
        tech_mental_data = cursor.fetchall()
        
        # 2. Screen Time vs Sleep Quality
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN screen_time_hours < 4 THEN 'Low Screen Time (< 4hrs)'
                    WHEN screen_time_hours < 8 THEN 'Medium Screen Time (4-8hrs)'
                    ELSE 'High Screen Time (> 8hrs)'
                END as screen_time_level,
                ROUND(AVG(sleep_hours), 2) as avg_sleep_hours,
                COUNT(*) as sample_size
            FROM fy_group_data
            WHERE screen_time_hours IS NOT NULL
            AND sleep_hours IS NOT NULL
            GROUP BY 
                CASE 
                    WHEN screen_time_hours < 4 THEN 'Low Screen Time (< 4hrs)'
                    WHEN screen_time_hours < 8 THEN 'Medium Screen Time (4-8hrs)'
                    ELSE 'High Screen Time (> 8hrs)'
                END
            ORDER BY avg_sleep_hours DESC
        """)
        screen_sleep_data = cursor.fetchall()
        
        # 3. Social Media Usage vs Stress Level
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN social_media_usage_hours < 2 THEN '< 2hrs'
                    WHEN social_media_usage_hours < 4 THEN '2-4hrs'
                    WHEN social_media_usage_hours < 6 THEN '4-6hrs'
                    ELSE '> 6hrs'
                END as social_media_usage,
                stress_level,
                COUNT(*) as count
            FROM fy_group_data
            WHERE social_media_usage_hours IS NOT NULL
            AND stress_level IS NOT NULL
            GROUP BY 
                CASE 
                    WHEN social_media_usage_hours < 2 THEN '< 2hrs'
                    WHEN social_media_usage_hours < 4 THEN '2-4hrs'
                    WHEN social_media_usage_hours < 6 THEN '4-6hrs'
                    ELSE '> 6hrs'
                END,
                stress_level
            ORDER BY social_media_usage
        """)
        social_stress_data = cursor.fetchall()
        
        # 4. Gaming Hours vs Physical Activity
        cursor.execute("""
            SELECT 
                CASE 
                    WHEN gaming_hours < 2 THEN 'Light Gamer (< 2hrs)'
                    WHEN gaming_hours < 4 THEN 'Moderate Gamer (2-4hrs)'
                    ELSE 'Heavy Gamer (> 4hrs)'
                END as gaming_level,
                ROUND(AVG(physical_activity_hours), 2) as avg_physical_activity,
                COUNT(*) as sample_size
            FROM fy_group_data
            WHERE gaming_hours IS NOT NULL
            AND physical_activity_hours IS NOT NULL
            GROUP BY 
                CASE 
                    WHEN gaming_hours < 2 THEN 'Light Gamer (< 2hrs)'
                    WHEN gaming_hours < 4 THEN 'Moderate Gamer (2-4hrs)'
                    ELSE 'Heavy Gamer (> 4hrs)'
                END
            ORDER BY avg_physical_activity DESC
        """)
        gaming_physical_data = cursor.fetchall()

    # Prepare data for charts
    chart_data = {
        'tech_mental': {
            'labels': [row[0] for row in tech_mental_data],
            'data': [row[1] for row in tech_mental_data],
            'percentages': [float(row[2]) for row in tech_mental_data]
        },
        'screen_sleep': {
            'labels': [row[0] for row in screen_sleep_data],
            'sleep_hours': [float(row[1]) for row in screen_sleep_data],
            'sample_sizes': [row[2] for row in screen_sleep_data]
        },
        'social_stress': {
            'labels': list(set([row[0] for row in social_stress_data])),
            'stress_levels': list(set([row[1] for row in social_stress_data])),
            'datasets': [[row[2] for row in social_stress_data if row[1] == level] 
                        for level in set([row[1] for row in social_stress_data])]
        },
        'gaming_physical': {
            'labels': [row[0] for row in gaming_physical_data],
            'physical_hours': [float(row[1]) for row in gaming_physical_data],
            'sample_sizes': [row[2] for row in gaming_physical_data]
        }
    }

    return render_template('dashboard.html', chart_data=chart_data)

# Clear Data ############################################################
@app.route('/group_clear', methods=['POST'])
def group_clear():
    with OracleDB().get_connection() as connection:
        cursor = connection.cursor()
        try:
            # Clear all data
            sql = "TRUNCATE TABLE FY_GROUP_DATA"
            cursor.execute(sql)
            
            # Reset sequence
            try:
                cursor.execute("DROP SEQUENCE seq_fy_group_data")
            except:
                pass  # Sequence might not exist
            
            cursor.execute("""
                CREATE SEQUENCE seq_fy_group_data
                START WITH 1
                INCREMENT BY 1
                NOCACHE
                NOCYCLE
            """)
            
            connection.commit()
            flash('All records have been cleared and sequence has been reset successfully', 'success')
        except Exception as e:
            flash(f'Error clearing data: {str(e)}', 'danger')
            
    return redirect(url_for('group_data'))

# Reset Sequence ############################################################

@app.route('/reset_sequence', methods=['POST'])
def reset_sequence():
    with OracleDB().get_connection() as connection:
        cursor = connection.cursor()
        try:
            # Drop existing sequence if it exists
            try:
                cursor.execute("DROP SEQUENCE fy_group_data_seq")
            except:
                pass  # Sequence might not exist
            
            # Create new sequence starting at 1
            cursor.execute("""
                CREATE SEQUENCE fy_group_data_seq
                START WITH 1
                INCREMENT BY 1
                NOCACHE
                NOCYCLE
            """)
            connection.commit()
            flash('Sequence has been reset successfully', 'success')
        except Exception as e:
            flash(f'Error resetting sequence: {str(e)}', 'danger')
    return redirect(url_for('group_data'))

@app.route('/check_data')
def check_data():
    with OracleDB().get_connection() as connection:
        cursor = connection.cursor()
        cursor.execute("SELECT COUNT(*) FROM FY_GROUP_DATA")
        count = cursor.fetchone()[0]
        return jsonify({'has_data': count > 0})

######################################################################

if __name__ == '__main__':
    app.run(debug=True)