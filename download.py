from database import OracleDB
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import Reference, BarChart
import pandas as pd

# Create a new workbook and activate the first worksheet
wb = Workbook()
ws = wb.active

with OracleDB().get_connection() as connection:
    query = "SELECT * FROM fy_group_data"
    cursor = connection.cursor()
    cursor.execute(query)
    data = cursor.fetchall()


#headers
row = 1
ws.cell(row=row,column=1).value = "Data ID"
ws.cell(row=row,column=2).value = "User ID"
ws.cell(row=row,column=3).value = "Age"
ws.cell(row=row,column=4).value = "Gender"
ws.cell(row=row,column=5).value = "Technology Usage Hours"
ws.cell(row=row,column=6).value = "Social Media Usage Hours"
ws.cell(row=row,column=7).value = "Gaming Hours"
ws.cell(row=row,column=8).value = "Screen Time Hours"
ws.cell(row=row,column=9).value = "Mental Health Status"
ws.cell(row=row,column=10).value = "Stress Level"
ws.cell(row=row,column=11).value = "Sleep Hours"
ws.cell(row=row,column=12).value = "Physical Activity Hours"
ws.cell(row=row,column=13).value = "Support Systems Access"
ws.cell(row=row,column=14).value = "Work Environment Impact"
ws.cell(row=row,column=15).value = "Online Support Usage"


for col in range(1,16):
    ws.cell(row=row,column=col).style = "Accent1"

row = 2

#data
for data_row in data:
    ws.cell(row=row,column=1).value = data_row[0]
    ws.cell(row=row,column=2).value = data_row[1]
    ws.cell(row=row,column=3).value = data_row[2]
    ws.cell(row=row,column=4).value = data_row[3]
    ws.cell(row=row,column=5).value = data_row[4]
    ws.cell(row=row,column=6).value = data_row[5]
    ws.cell(row=row,column=7).value = data_row[6]
    ws.cell(row=row,column=8).value = data_row[7]
    ws.cell(row=row,column=9).value = data_row[8]
    ws.cell(row=row,column=10).value = data_row[9]
    ws.cell(row=row,column=11).value = data_row[10]
    ws.cell(row=row,column=12).value = data_row[11]
    ws.cell(row=row,column=13).value = data_row[12]
    ws.cell(row=row,column=14).value = data_row[13]
    ws.cell(row=row,column=15).value = data_row[14]
    row += 1

#expand all columns to fit the data
for i in range(1,ws.max_column+1):
    ws.column_dimensions[get_column_letter(i)].bestFit = True
    ws.column_dimensions[get_column_letter(i)].auto_size = True


# Save the workbook to a new copy
wb.save('new_data.xlsx')

